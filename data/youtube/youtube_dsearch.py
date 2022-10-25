# Imports
from youtube_search import YoutubeSearch
import pandas as pd
import openpyxl 
import networkx as nx
import matplotlib.pyplot as plt
import urllib.parse as urp
from googleapiclient.discovery import build 

# Define main variables
# API credentials https://console.developers.google.com/
API_KEY = open('youtube/API_KEY.txt').readline() 

#build a resource for youtube
resource = build('youtube', 'v3', developerKey=API_KEY) 
input_data = 'youtube/youtube.xlsx'
output_data = 'youtube/youtube_output.xlsx'
maxim_results = 10000
max_comments = 50
search_terms_list = ['distance and senegal','Covid-19 and senegal','fomites and senegal',
'pandemic and senegal','spreading and senegal','quarantine and senegal','isolation and senegal',
'contagious and senegal','coronavirus and senegal', 'epidemic and senegal']


def main(input_data= input_data, output_data= output_data, 
maxim_results = maxim_results, max_comments=max_comments, 
search_terms_list= search_terms_list):  

    # Load columns data
    workbook = openpyxl.load_workbook(input_data)
    sheet = workbook["Sheet1"]
    max_rows = sheet.max_row
    max_column = sheet.max_column

    #Define a dictionary to hold the column names
    ColNames={}
    Current=1

    for Col in sheet.iter_cols(1, sheet.max_column):
        ColNames[Col[0].value]=Current
        Current +=1 

    # Are these the required key words?
    #search_terms_list = ['no poverty','zero hunger', 'good health and well-being','quality education', 'gender equality']
    #search_terms_list = ['poverty and senegal','destitute and senegal','bankrupthy and senegal','impecunious and senegal','lack and senegal','shortage and senegal','scarcity and senegal','necessity and senegal','poor and senegal']
    #search_terms_list = ['hunger and senegal','craving and senegal', 'desire and senegal','famine and senegal','greed and senegal','longing and senegal','lust and senegal','starvation and senegal','yearning and senegal','food and senegal']

    # How many results are required?
    

    results_dict = {}

    row_count =2

   # node_names = []
   #edges = []

    for item in search_terms_list:
        results = YoutubeSearch(item, max_results=maxim_results).to_dict()
        results_dict[item] = results
        for video in range(len(results_dict[item])):
            try:                    
                # Get the comments under this video - url suffix 
                video_id = str(results_dict[item][video]['url_suffix'])[9:] 
                video_id = urp.unquote(video_id,encoding='utf-8', errors='replace') 
                #create a request to get  comments on the video
                request = resource.commentThreads().list(
                                            part="snippet",
                                            videoId=video_id,
                                            maxResults= max_comments,   #get  comments
                                            order="orderUnspecified")  #top comments.
                #execute the request
                response =request.execute()

                #get first
                r_items = response["items"]

                for it in r_items:
                    item_info = it["snippet"]
                    
                    sheet.cell(row=row_count, column= ColNames['keyword']).value = item
                   # node_names.append(item)
                    sheet.cell(row=row_count, column= ColNames['title']).value = results_dict[item][video]['title']
                    #edges.append((item, results_dict[item][video]['title']))
                    sheet.cell(row=row_count, column= ColNames['channel']).value = results_dict[item][video]['channel']
                    sheet.cell(row=row_count, column= ColNames['duration']).value = results_dict[item][video]['duration']
                    sheet.cell(row=row_count, column= ColNames['views']).value = results_dict[item][video]['views']
                    sheet.cell(row=row_count, column= ColNames['id']).value = results_dict[item][video]['id']
                    sheet.cell(row=row_count, column= ColNames['publish_time']).value = results_dict[item][video]['publish_time']
                    sheet.cell(row=row_count, column= ColNames['thumbnails']).value = ', '.join(results_dict[item][video]['thumbnails']) 
                    sheet.cell(row=row_count, column= ColNames['url_suffix']).value = results_dict[item][video]['url_suffix']

                    #the top level comment can have sub reply comments
                    topLevelComment = item_info["topLevelComment"]
                    comment_info = topLevelComment["snippet"]
                                
                    sheet.cell(row=row_count, column= ColNames['cmt_author']).value = comment_info["authorDisplayName"]
                    sheet.cell(row=row_count, column= ColNames['comment']).value =  comment_info["textDisplay"]
                    sheet.cell(row=row_count, column= ColNames['cmt_likes']).value = comment_info["likeCount"]
                    sheet.cell(row=row_count, column= ColNames['cmt_time']).value = comment_info['publishedAt']
                    print(comment_info["textDisplay"])
                    row_count += 1
            except Exception as e:
                # Some URLs require URL encoding before being passed into the build resource
                pass

    workbook.save(output_data)

#def draw_netgraph(edges, node_names):
    #print('Edges',len(edges))
    #print('Nodes',len(node_names))

    # What variables are you interested in to put into the graph
   # G = nx.Graph()
   # G.add_nodes_from(node_names) # Add nodes to the Graph
   # G.add_edges_from(edges) # Add edges to the Graph
    #print(nx.info(G)) # Print information about the Graph

   # nx.draw(G) 
   # plt.savefig("Network_graph.png") # save as png
   # plt.show() # display

    # Requirement: what metrics are required - centrality
    # Interface for the application domain 

# Call the main function 
main()