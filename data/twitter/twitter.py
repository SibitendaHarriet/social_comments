# Import the required modules
import os
from unicodedata import name
import requests
import json
from nltk.tokenize import sent_tokenize, word_tokenize
import matplotlib.pyplot as plt
from nltk.probability import FreqDist
import openpyxl
import pandas as pd
import time

BEARER_TOKEN = open('twitter/bearer.txt').readline()
#search_terms_list = ['distance and senegal','Covid-19 and senegal','fomites and senegal','pandemic and senegal','spreading and senegal','quarantine and senegal','isolation and senegal','contagious and senegal','coronavirus and senegal', 'epidemic and senegal']

search_terms_list = ['distance 'and' senegal','Covid-19 'and' senegal','fomites 'and' senegal','pandemic 'and' senegal','spreading 'and' senegal','quarantine 'and' senegal','isolation 'and' senegal','contagious 'and' senegal','coronavirus 'and' senegal', 'epidemic 'and' senegal']
input_data = 'twitter/twitter.xlsx'
output_data = 'twitter/twitter_output.xlsx'

# define search twitter function
def search_twitter(query, tweet_fields, bearer_token=BEARER_TOKEN):
    headers = {"Authorization": "Bearer {}".format(bearer_token)}

    url = "https://api.twitter.com/2/tweets/search/recent?max_results=100&query={}&{}".format(
        query, tweet_fields)
    response = requests.request("GET", url, headers=headers)

    print('Response status code:',response.status_code)

    if response.status_code != 200:
        raise Exception(response.status_code, response.text)
    return response.json()

def main(input_data = input_data, output_data = output_data, search_terms_list = search_terms_list):

    # Load input workbook to get dataset columns 
    workbook = openpyxl.load_workbook(input_data)
    sheet = workbook["Sheet1"]
    max_rows = sheet.max_row
    max_column = sheet.max_column

    # Define a dictionary to hold the column names
    ColNames = {}
    Current = 1
    for Col in sheet.iter_cols(1, sheet.max_column):
        ColNames[Col[0].value] = Current
        Current += 1

    # Which API response fields are we interested in?
    tweet_fields = "tweet.fields=text,author_id,created_at,conversation_id,entities,geo,id,in_reply_to_user_id,lang,public_metrics,possibly_sensitive,source"

    # Initialize row count 
    filled_row_count = 2 

    # Search through all terms in the search terms list
    for key_word in search_terms_list:
        # Call the twitter API 
        json_response = search_twitter(
            query=key_word, tweet_fields=tweet_fields, bearer_token=BEARER_TOKEN) 

        # Store the returned data, catering for all the possible outcomes
        try:
            for each_tweet in range(len(json_response['data'])):
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['keyword']).value = key_word
                except Exception as e:
                    pass        
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['result_id']).value = each_tweet
                except Exception as e:
                    pass        
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['cum_result_id']).value = each_tweet
                except Exception as e:
                    pass        
                try:
                    sheet.cell(row=filled_row_count, column=ColNames['created_at']
                            ).value = json_response['data'][each_tweet]['created_at']
                except Exception as e:
                    pass

                    sheet.cell(row=filled_row_count,
                            column=ColNames['id']).value = json_response['data'][each_tweet]['id']
                except Exception as e:
                    pass
                try:    
                    sheet.cell(row=filled_row_count,
                            column=ColNames['author_id']).value = json_response['data'][each_tweet]['author_id']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['text']).value = json_response['data'][each_tweet]['text']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['newest_id']).value = json_response['meta']['newest_id']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['oldest_id']).value = json_response['meta']['oldest_id']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['result_count']).value = json_response['meta']['result_count']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['next_token']).value = json_response['meta']['next_token']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count, column=ColNames['conversation_id']
                            ).value = json_response['data'][each_tweet]['conversation_id']
                except Exception as e:
                    pass

                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['geo']).value = json_response['data'][each_tweet]['geo']
                except Exception as e:
                    pass

                try:
                    sheet.cell(row=filled_row_count, column=ColNames['in_reply_to_user_id']
                            ).value = json_response['data'][each_tweet]['in_reply_to_user_id']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count,
                            column=ColNames['lang']).value = json_response['data'][each_tweet]['lang']
                except Exception as e:
                    pass                
                try:
                    sheet.cell(row=filled_row_count, column=ColNames['retweet_count']
                            ).value = json_response['data'][each_tweet]['public_metrics']['retweet_count']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count, column=ColNames['reply_count']
                            ).value = json_response['data'][each_tweet]['public_metrics']['reply_count']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count, column=ColNames['like_count']
                            ).value = json_response['data'][each_tweet]['public_metrics']['like_count']
                except Exception as e:
                    pass
                try:
                    sheet.cell(row=filled_row_count, column=ColNames['quote_count']
                            ).value = json_response['data'][each_tweet]['public_metrics']['quote_count']
                except Exception as e:
                    pass
                try:
                    users = ''
                    for user in range(len(json_response['data'][each_tweet]['entities']['mentions'])):
                        users += json_response['data'][each_tweet]['entities']['mentions'][user]['username']
                        users += ','
                    users = users[:-1]
                    sheet.cell(row=filled_row_count,
                            column=ColNames['mentions']).value = users
                except Exception as e:
                    pass
                try:
                    users = ''
                    for user in range(len(json_response['data'][each_tweet]['entities']['annotations'])):
                        users += json_response['data'][each_tweet]['entities']['annotations'][user]['normalized_text']
                        users += ','
                    users = users[:-1]
                    sheet.cell(row=filled_row_count, column=ColNames['annotations']).value = users
                except Exception as e:
                    pass
                try:
                    users = ''
                    for user in range(len(json_response['data'][each_tweet]['entities']['hashtags'])):
                        users += json_response['data'][each_tweet]['entities']['hashtags'][user]['tag']
                        users += ','
                    users = users[:-1] 
                    sheet.cell(row=filled_row_count, column=ColNames['hashtags']).value = users

                except Exception as e:
                    pass
                try:
                    users = ''
                    for user in range(len(json_response['data'][each_tweet]['entities']['urls'])):
                        users += json_response['data'][each_tweet]['entities']['urls'][user]['url']
                        users += ','
                    users = users[:-1]
                    # There can be many mentions
                    sheet.cell(row=filled_row_count,
                            column=ColNames['url']).value = users
                except Exception as e:
                    pass
                try:
                    users = ''
                    for user in range(len(json_response['data'][each_tweet]['entities']['urls'])):
                        users += json_response['data'][each_tweet]['entities']['urls'][user]['expanded_url']
                        users += ','
                    users = users[:-1]
                    sheet.cell(row=filled_row_count, column=ColNames['expanded_url']).value = users
                except Exception as e:
                    pass 
                try:
                    users = ''
                    for user in range(len(json_response['data'][each_tweet]['entities']['urls'])):
                        users += json_response['data'][each_tweet]['entities']['urls'][user]['display_url']
                        users += ','
                    users = users[:-1]
                    sheet.cell(row=filled_row_count, column=ColNames['display_url']).value = users
                except Exception as e:
                    pass

                filled_row_count += 1
        except:
            print('Empty response received, skipping ...')
            
        time.sleep(20) # Wait for some time to prevent overloading API server and subsequent blockage of access
    workbook.save(output_data) 
    return "success"

# Call the main function
main() 