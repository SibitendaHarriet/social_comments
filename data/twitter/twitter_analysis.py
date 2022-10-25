# Import the required modules
import os
import requests
import json
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.probability import FreqDist
from nltk.corpus import stopwords 
from nltk.stem import PorterStemmer 
from nltk.stem import WordNetLemmatizer
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import numpy as np 
import time
import networkx as nx 

'''
import re
import string

import gensim
from gensim import corpora

# libraries for visualization
import pyLDAvis
import pyLDAvis.gensim
import seaborn as sns
'''

#input_data = r'E:\work\silicon savannah\repositories\pyrepo\twitter.xlsx'
input_data = 'twitter/twitter.xlsx'
workbook = openpyxl.load_workbook(input_data)
sheet = workbook["Sheet1"]
max_rows = sheet.max_row
max_column = sheet.max_column

# Define a dictionary to hold the column names
ColNames = {}
Current = 1

for Col in sheet.iter_cols(1, sheet.max_column):
    ColNames[Col[0].value] = Current
    Current += 1

# Do NLP transformation
# Keywords in the text 
# Sentence Tokenization 
# Removal of stop words 
# Stemming and lemmatization 

lemmatizer = WordNetLemmatizer()

node_names = []    # Define nodes
edges = [] # Define edges

all_data = []
all_dataset = [] 

# What variables are you interested in to put into the graph
G = nx.Graph()

for tweet in range(2, 45) : #max_rows+1
    if sheet.cell(row=tweet, column =ColNames['lang']).value =='en':
        tweet_text = sheet.cell(row=tweet, column =ColNames['text']).value
        tokenized_text=sent_tokenize(tweet_text)
        tokenized_word=word_tokenize(tweet_text)

        fdist = FreqDist(tokenized_word)
        # Frequency Distribution Plot
        #fdist.plot(30,cumulative=False)
        #plt.show()

        # Remove stop words 
        tokenized_word_wsw = [word for word in tokenized_word if not word in stopwords.words()]

        # Stemming will not be applied, apply lemmatization
        tokenized_lemmas = [lemmatizer.lemmatize(word) for word in tokenized_word_wsw] 
        all_data.append(tokenized_lemmas) 
        for item in tokenized_lemmas:
            all_dataset.append(item) 

all_dataset = list(set(all_dataset)) 
print(all_dataset)

'''
dictionary = corpora.Dictionary(all_dataset)
doc_term_matrix = [dictionary.doc2bow(rev) for rev in all_dataset] 

print(dictionary)
print(doc_term_matrix)


# Creating the object for LDA model using gensim library
LDA = gensim.models.ldamodel.LdaModel

# Build LDA model
lda_model = LDA(corpus=doc_term_matrix, id2word=dictionary, num_topics=10, random_state=100,
                chunksize=1000, passes=50,iterations=100)

lda_model.print_topics()


# Visualize the topics
#https://github.com/bmabey/pyLDAvis
#https://speakerdeck.com/bmabey/visualizing-topic-models
#pyLDAvis.enable_notebook()
#vis = pyLDAvis.gensim.prepare(lda_model, doc_term_matrix, dictionary)
#vis


print('\nPerplexity: ', lda_model.log_perplexity(doc_term_matrix,total_docs=10000))  # a measure of how good the model is. lower the better.

# Compute Coherence Score
from gensim.models.coherencemodel import CoherenceModel
coherence_model_lda = CoherenceModel(model=lda_model, texts=all_dataset, dictionary=dictionary , coherence='c_v')
coherence_lda = coherence_model_lda.get_coherence()
print('\nCoherence Score: ', coherence_lda)
'''
tweet_count =0
for each_tweet in all_data:
    try:
        tweet_count +=1 
        for each_word in each_tweet:
            for other_word in each_tweet:
                if each_word!= other_word:
                    G.add_edge(each_word, other_word, label= tweet_count)

    except Exception as e:
        print(e) 

pos = nx.spring_layout(G)
edge_labels = nx.get_edge_attributes(G, 'label')
nx.draw(G, pos)
nx.draw_networkx_edge_labels(G, pos, edge_labels, font_size=7)
nx.draw_networkx_labels(G, pos, font_size=10)

plt.savefig("twitter_graph.png") # save as png
plt.show() # display

# Requirement: what metrics are required - centrality
# Interface for the application domain 

# https://stackabuse.com/python-for-nlp-topic-modeling/ 

'''

Pre-requisites
gensim: https://radimrehurek.com/gensim/
pyLDAvis : https://pyldavis.readthedocs.io/en/latest/readme.html#installation
matplotlib: https://matplotlib.org/stable/users/installing.html
spacy 2.0 : https://pypi.org/project/spacy/2.0.0/
LDA paper :
https://ai.stanford.edu/~ang/papers/jair03-lda.pdf
LDA tutorial:
https://en.wikipedia.org/wiki/Latent_Dirichlet_allocation
https://www.machinelearningplus.com/nlp/topic-modeling-gensim-python/
'''
